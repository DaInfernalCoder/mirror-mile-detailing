#!/usr/bin/env python3
"""Upsert one make-gmb business record into the existing Profiles workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("record", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    record = json.loads(args.record.read_text())
    workbook = load_workbook(args.workbook)
    if workbook.sheetnames != ["Profiles"]:
        raise SystemExit(f"Unexpected sheets: {workbook.sheetnames}")

    sheet = workbook["Profiles"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if len(headers) != 41 or "canonical_business_name" not in headers:
        raise SystemExit("Registry schema does not match the expected 41-column Profiles layout")

    canonical_column = headers.index("canonical_business_name") + 1
    canonical_name = record["canonical_business_name"]
    target_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, canonical_column).value == canonical_name:
            target_row = row
            break
    if target_row is None:
        target_row = sheet.max_row + 1

    wrap_fields = {
        "gbp_description",
        "services",
        "missing_items",
        "notes",
        "drive_asset_folder",
    }
    for column, header in enumerate(headers, start=1):
        value = record.get(header, "")
        cell = sheet.cell(target_row, column)
        cell.value = value
        if header in wrap_fields:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )
        if header in {"phone", "domain_auto_renew"}:
            cell.number_format = "@"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:AO{sheet.max_row}"
    workbook.save(args.workbook)

    check = load_workbook(args.workbook, read_only=False, data_only=False)
    verify = check["Profiles"]
    matches = [
        row
        for row in range(2, verify.max_row + 1)
        if verify.cell(row, canonical_column).value == canonical_name
    ]
    if len(matches) != 1 or verify.max_column != 41:
        raise SystemExit("Registry verification failed after save")

    print(
        json.dumps(
            {
                "row": matches[0],
                "rows": verify.max_row,
                "columns": verify.max_column,
                "freeze": str(verify.freeze_panes),
                "filter": verify.auto_filter.ref,
                "canonical_business_name": canonical_name,
                "setup_status": verify.cell(matches[0], headers.index("setup_status") + 1).value,
            }
        )
    )


if __name__ == "__main__":
    main()
